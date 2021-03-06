from collections import defaultdict
from typing import Iterable

import pandas as pd
from pandas import DataFrame, MultiIndex
from openpyxl import Workbook
from datetime import datetime, timedelta


def parse_pegas(file_name):
    df = pd.read_excel(file_name, sheet_name='Dannie')

    df = df.rename(columns={'Дата включения': 'date',
                            'Код филиала': 'branch_code',
                            'Причина приостановки уплаты взносов': 'cancellation_reason'})

    return df


def parse_drfs(file_name: str) -> DataFrame:
    df = pd.read_excel(file_name, sheet_name='Лист1')

    df = df.reset_index(level=[1, 2], drop=True).dropna().reset_index()

    df = df.dropna().set_index('index')
    df.index = df.index.map(int)

    return df


def is_supporting_date(dt: datetime):
    supporting_months = [2, 5, 8, 11]
    if dt.month in supporting_months:
        if dt.day == 1:
            return True
    return False


def date_range(dt: datetime, days: int) -> Iterable[datetime]:
    for i in range(days):
        yield dt
        dt = dt + timedelta(days=1)


def save_to_excel_pegas_data(data: defaultdict) -> None:
    wb = Workbook()
    for branche_code, supporting_dates in data.items():

        ws = wb.create_sheet(title=str(branche_code))
        ws.merge_cells("A1:C1")
        ws["A1"].value = "Для файла Пегас"
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

        ws["A2"] = "Начало периода"
        ws["B2"] = "Конец периода"
        ws["C2"] = "Количество дней"

        cell_number = 3
        day_range = 20
        period = 60
        for support_date in supporting_dates:
            for dt in date_range(support_date, day_range):
                ws[f"A{cell_number}"].value = (dt - timedelta(days=period)).strftime("%Y-%m-%d")
                ws[f"B{cell_number}"].value = dt.strftime("%Y-%m-%d")
                ws[f"C{cell_number}"].value = period
                cell_number += 1

            cell_number += 1

    wb.save("test.xlsx")


def calc_pegas_periods(data: DataFrame):
    pegas_branch_to_date_map = defaultdict(set)
    for branch_code, group in data.groupby(['branch_code']):
        for i, row in group.iterrows():
            pegas_branch_to_date_map[branch_code].add(row['date'])

    branch_codes = [branch_code for branch_code, supporting_dates in pegas_branch_to_date_map.items()]
    multi_index = MultiIndex.from_product([branch_codes, ['period_start', 'period_end', 'population']],
                                          names=['branch_name', 'values'])

    df = DataFrame([[]], index=multi_index)

    day_range = 20
    for branch_code, supporting_dates in data.items():
        for support_date in supporting_dates:
            for dt in date_range(support_date, day_range):
                df = df.loc['branch_code', 'period_end'] = [branch_code, dt]


    # save_to_excel_pegas_data(pegas_branch_to_date_map)


def form_forecast_report(fact_data, pegas_data, drfs_data):
    pass


if __name__ == "__main__":
    pegas_file_name = "dataset_pegas_test.xlsx"
    drfs_file_name = "dataset_drfs_test.xlsx"
    fact_file_name = "dataset_fact_test.xlsx"

    pegas_raw_data = parse_pegas(pegas_file_name)
    pegas_data = calc_pegas_periods(pegas_raw_data)

    # drfs_data = parse_drfs(drfs_file_name)
